from django.shortcuts import render
from django.http import HttpResponse
from .forms import CollectForm
from .getdata import getdata
from .dataDao import dataDao
from .analytics import analytics
from .models import data
from openpyxl import Workbook
import openpyxl
import base64
# =============================================================================
# from flask import Flask, session
# app = Flask(__name__)
# app.secret_key = "secret"
# =============================================================================

# Create your views here.
def index(request):
    params = {
              'title': 'Hello',
              'message': 'yourdata:',
              'form': CollectForm()
        }
    if (request.method == 'POST'):
        params['message'] = '名前:'+request.POST['msg']
        params['form'] = CollectForm(request.POST)
    return render(request, 'collect/index.html',params)

#urlを入力したフォームが提出されたときの処理
#@app.route("/collect")
def form(request):
# =============================================================================
#     url = request.POST['url']
#     date = request.POST['date']
#     storeName = request.POST['storeName']
# =============================================================================
    BUTTON_ID = request.POST['BUTTON_ID']
    # getdataクラスのインスタンスを作成
    gd = getdata()
    #data_arr = gd.culcurate(url)
    dao = dataDao()
    #anallyticsクラスのインスタンスを作成
    ana = analytics()


    if BUTTON_ID == "001":
        # culcurateメソッドを呼び出し、URLからデータを取得する
        url = request.POST['url']
        date = request.POST['date']
        storeName = request.POST['storeName']
        #すでにデータベースにデータがないかチェック
        check_data = dao.check_data(date, storeName)
        if not check_data: 
            # getdataクラスのインスタンスを作成
            gd = getdata()
            data_arr = gd.culcurate(url)
            params = {
                      'title':'Hello/Form',
                      'url':url,
                      'data':data_arr
                }
            #データベースにデータを挿入
            dao.insert_data(date, storeName, data_arr)
    
            #sessionオブジェクトにurlから持ってきたデータ配列の格納
            request.session["data"] = data_arr
            #session.save()
            return render(request, 'collect/result.html',params)
        else:
            params = {
                      'error':'すでにデータが存在しています',
                }
            return render(request, 'collect/index.html',params)
    
    
    if BUTTON_ID == "002":
# =============================================================================
#         if 'data' in request.session:
#             data2 = request.session["data"]
#             wb = Workbook() # ワークブックオブジェクトを作成
#             ws = wb.active # アクティブなワークシートオブジェクトを取得
#             for row in data2: # 配列のデータをループで処理
#                 ws.append(row) # ワークシートに行を追加
#             wb.save("test.xlsx") # ファイル名を指定して保存
# =============================================================================
            selectdata = dao.test_data("マルハン新宿東方")
            params = {
                'selectdata' : selectdata,
                'x' : selectdata[7].modelName
                }
            return render(request, 'collect/ex.html',params)
        
    if BUTTON_ID == "003":
        date1 = request.POST['date1']
        date2 = request.POST['date2']
        
        model,mse = ana.regression_analysis(date1, date2)
        # dataをbase64エンコードして文字列に変換
        #img = base64.b64encode(data).decode()
        
        params = {
                  'model':model,
                  'mse':mse
            }
        
        return render(request, 'collect/ana.html',params)
        
# =============================================================================
#     #失敗　既存のエクセルファイル更新
#     if BUTTON_ID == "002":
#         excelfile = request.POST['excelfile']
#         if 'data' in request.session:
#             data2 = request.session["data"]
#             # 既存のExcelファイルを読み込む
#             wb = openpyxl.load_workbook(excelfile)
#             # 新しいシートを作成する
#             ws = wb.create_sheet("new_sheet")
#             for row in data2: # 配列のデータをループで処理
#                 ws.append(row) # ワークシートに行を追加
#             wb.save(excelfile) # ファイル名を指定して保存
#             
#             return render(request, 'collect/result.html')
# =============================================================================

# =============================================================================
# #urlを入力したフォームが提出されたときの処理
# def excel(request):
#             wb = Workbook() # ワークブックオブジェクトを作成
#             ws = wb.active # アクティブなワークシートオブジェクトを取得
#             for row in data: # 配列のデータをループで処理
#                 ws.append(row) # ワークシートに行を追加
#             wb.save("test.xlsx") # ファイル名を指定して保存
#             
#             return render(request, 'collect/result.html')
# =============================================================================
# =============================================================================
#データベースに格納 daoにメソッド作成
#         for i in range(1,len(data_arr)):
#             if len(data_arr[i])==9:
#                 obj = data(storeName=storeName,
#                        date=date,
#                        modelName=data_arr[i][0],
#                        number = data_arr[i][1],
#                        game = data_arr[i][2],
#                        difference = data_arr[i][3],
#                        BB = data_arr[i][4],
#                        RB = data_arr[i][5],
#                        allprobability = data_arr[i][6],
#                        BBprobability = data_arr[i][7],
#                        RBprobability = data_arr[i][8])
#                 obj.save()
# =============================================================================

